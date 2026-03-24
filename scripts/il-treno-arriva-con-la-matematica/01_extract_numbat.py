#!/usr/bin/env python3
"""
01_extract_numbat.py — Estrai dati NUMBAT (TfL Londra)
Legge il file Excel NUMBAT e produce CSV puliti:
  - station_entries_15min.csv  (passeggeri per stazione ogni 15 min)
  - link_frequencies.csv       (treni per link per fascia oraria)
  - station_rho.csv            (rho = lambda/mu per stazione)

Fonte: TfL NUMBAT 2024 (Tue-Wed-Thu, giorno tipo feriale)
"""

import openpyxl
import csv
import os

NUMBAT_FILE = 'data/numbat_twt.xlsx'
OUTPUT_DIR = 'data'

# Fasce orarie NUMBAT (15-min slots da 0500 a 0100)
TIME_SLOTS = []
for h in range(5, 25):  # 5:00 -> 24:00
    hh = h % 24
    for q in range(4):
        start = f'{hh:02d}{q*15:02d}'
        end_m = (q + 1) * 15
        end_h = hh
        if end_m == 60:
            end_m = 0
            end_h = (hh + 1) % 24
        end = f'{end_h:02d}{end_m:02d}'
        TIME_SLOTS.append(f'{start}-{end}')
# Aggiunge 0100-0115, 0115-0130 se presenti
for q in range(4):
    hh = 1
    start = f'{hh:02d}{q*15:02d}'
    end_m = (q + 1) * 15
    end_h = hh
    if end_m == 60:
        end_m = 0
        end_h = 2
    end = f'{end_h:02d}{end_m:02d}'
    slot = f'{start}-{end}'
    if slot not in TIME_SLOTS:
        TIME_SLOTS.append(slot)


def extract_station_entries():
    """Estrai ingressi per stazione a 15 min."""
    print('[*] Apertura NUMBAT...')
    wb = openpyxl.load_workbook(NUMBAT_FILE, read_only=True)
    ws = wb['Station_Entries']

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Riga 2 (index 2) = header con nomi colonne
    header = rows[2]
    # Trova indice colonna '0500-0515'
    col_start = None
    for i, v in enumerate(header):
        if v and str(v).strip() == '0500-0515':
            col_start = i
            break

    if col_start is None:
        print('[!] Colonna 0500-0515 non trovata')
        return

    n_slots = len(header) - col_start
    print(f'[+] Trovate {n_slots} fasce orarie da 15 min')

    outfile = os.path.join(OUTPUT_DIR, 'station_entries_15min.csv')
    with open(outfile, 'w', newline='') as f:
        w = csv.writer(f)
        slots = [str(header[col_start + i]).replace(' ', '') if header[col_start + i] else f'slot_{i}'
                 for i in range(n_slots)]
        w.writerow(['station', 'fare_zone', 'total'] + slots)

        count = 0
        for row in rows[3:]:
            if not row[0] or not row[2]:
                continue
            station = str(row[2]).strip()
            zone = str(row[3]).strip() if row[3] else ''
            total = row[4] if row[4] else 0
            if total == 0:
                continue
            values = []
            for i in range(n_slots):
                idx = col_start + i
                v = row[idx] if idx < len(row) and row[idx] else 0
                values.append(round(float(v), 2))
            w.writerow([station, zone, round(float(total), 2)] + values)
            count += 1

    print(f'[+] {count} stazioni salvate in {outfile}')
    return outfile


def extract_link_frequencies():
    """Estrai frequenze treni per link."""
    wb = openpyxl.load_workbook(NUMBAT_FILE, read_only=True)
    ws = wb['Link_Frequencies']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = rows[2]

    outfile = os.path.join(OUTPUT_DIR, 'link_frequencies.csv')
    with open(outfile, 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['line', 'direction', 'from_station', 'to_station',
                     'total', 'early', 'am_peak', 'midday', 'pm_peak', 'evening', 'late'])

        count = 0
        for row in rows[3:]:
            if not row[0] or not row[1]:
                continue
            line = str(row[1]).strip()
            direction = str(row[2]).strip()
            from_st = str(row[6]).strip()
            to_st = str(row[9]).strip()
            vals = [row[i] if i < len(row) and row[i] else 0 for i in range(10, 17)]
            w.writerow([line, direction, from_st, to_st] + [round(float(v), 1) for v in vals])
            count += 1

    print(f'[+] {count} link salvati in {outfile}')
    return outfile


def compute_station_rho():
    """
    Calcola rho = lambda / mu per le stazioni principali.
    lambda = passeggeri / 15 min (peak)
    mu = treni / 15 min * capacita' treno (stimata)
    """
    import csv as csv_mod

    # Leggi entries
    entries = {}
    with open(os.path.join(OUTPUT_DIR, 'station_entries_15min.csv')) as f:
        reader = csv_mod.DictReader(f)
        fields = reader.fieldnames
        # Trova le colonne AM peak (0700-0930)
        peak_cols = [c for c in fields if c.startswith(('0700', '0715', '0730', '0745',
                                                         '0800', '0815', '0830', '0845',
                                                         '0900', '0915'))]
        for row in reader:
            station = row['station']
            peak_vals = [float(row[c]) for c in peak_cols if row[c]]
            if peak_vals:
                entries[station] = {
                    'total': float(row['total']),
                    'peak_avg_15min': sum(peak_vals) / len(peak_vals),
                    'peak_max_15min': max(peak_vals),
                    'zone': row['fare_zone']
                }

    # Leggi frequenze (treni per link, usiamo am_peak)
    freq = {}
    with open(os.path.join(OUTPUT_DIR, 'link_frequencies.csv')) as f:
        reader = csv_mod.DictReader(f)
        for row in reader:
            station = row['from_station']
            line = row['line']
            am = float(row['am_peak'])
            key = station
            if key not in freq:
                freq[key] = {'lines': set(), 'am_peak_trains': 0}
            freq[key]['lines'].add(line)
            freq[key]['am_peak_trains'] += am

    # AM Peak = 3h = 12 slot da 15min
    # freq e' il totale treni nell'AM Peak, dividiamo per 12 per avere treni/15min
    AM_PEAK_SLOTS = 12

    # Capacita' media treno LU = ~800 pax (standing + seated, 8 carrozze)
    TRAIN_CAPACITY = 800

    outfile = os.path.join(OUTPUT_DIR, 'station_rho.csv')
    with open(outfile, 'w', newline='') as f:
        w = csv_mod.writer(f)
        w.writerow(['station', 'zone', 'daily_total', 'lambda_peak_15min',
                     'trains_peak_15min', 'mu_peak_15min', 'rho', 'lines'])

        combined = []
        for station in entries:
            if station not in freq:
                continue
            e = entries[station]
            fr = freq[station]
            trains_15 = fr['am_peak_trains'] / AM_PEAK_SLOTS
            mu = trains_15 * TRAIN_CAPACITY
            lam = e['peak_max_15min']
            rho = lam / mu if mu > 0 else 0
            combined.append({
                'station': station,
                'zone': e['zone'],
                'total': e['total'],
                'lambda': round(lam, 1),
                'trains_15': round(trains_15, 2),
                'mu': round(mu, 1),
                'rho': round(rho, 4),
                'lines': '/'.join(sorted(fr['lines']))
            })

        combined.sort(key=lambda x: x['rho'], reverse=True)
        for c in combined:
            w.writerow([c['station'], c['zone'], c['total'], c['lambda'],
                        c['trains_15'], c['mu'], c['rho'], c['lines']])

    print(f'[+] rho calcolato per {len(combined)} stazioni -> {outfile}')
    top = combined[:10]
    print('\n[*] Top 10 stazioni per rho (AM Peak):')
    for c in top:
        print(f'    {c["station"]:30s}  rho={c["rho"]:.4f}  '
              f'lambda={c["lambda"]:.0f}  mu={c["mu"]:.0f}  ({c["lines"]})')


if __name__ == '__main__':
    extract_station_entries()
    extract_link_frequencies()
    compute_station_rho()
    print('\n[+] Done.')
